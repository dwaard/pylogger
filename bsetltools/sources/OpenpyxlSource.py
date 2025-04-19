import logging
import warnings
from openpyxl import load_workbook

class OpenpyxlSource:

    def __init__(self, file="", worksheet_index=0):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            workbook = load_workbook(filename = file, data_only=True)
        sheet_name = workbook.sheetnames[worksheet_index]
        logging.debug(f"Opening worksheet: {sheet_name}")
        self.worksheet = workbook[sheet_name]
        self.rows = self.worksheet.iter_rows(values_only=True)

    def __iter__(self):
        return self

    def __next__(self):
        while True:
            row = next(self.rows)  # Raise StopIteration als klaar
            if row is None or all(cell is None for cell in row):
                continue  # Sla lege rijen over

            trimmed = list(row)
            while trimmed and trimmed[-1] is None:
                trimmed.pop()
            
            return trimmed
